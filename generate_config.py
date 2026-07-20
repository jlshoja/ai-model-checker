#!/usr/bin/env python3
"""
generate_config.py
===================

Generates .env and providers.json files from providers.xlsx.

Usage:
    python generate_config.py
    python generate_config.py --output-dir ./config
    python generate_config.py --env-file .env.custom --providers providers_custom.json
"""

import argparse
import json
import os
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def sanitize_name(name: str) -> str:
    """Sanitize provider name for use as environment variable or JSON key."""
    # Remove parentheses and their contents
    name = re.sub(r'\(.*?\)', '', name)
    # Remove special characters and replace spaces with underscores
    name = re.sub(r'[^a-zA-Z0-9\s\-]', '', name)
    # Replace spaces and hyphens with underscores
    name = name.replace(' ', '_').replace('-', '_')
    # Replace multiple consecutive underscores with single underscore
    name = re.sub(r'_+', '_', name)
    # Remove leading/trailing underscores
    name = name.strip('_')
    return name


def read_providers_from_excel(xlsx_path: Path) -> list[dict]:
    """Read provider data from Excel file."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to read Excel files. Install with: pip install openpyxl")
    
    wb = load_workbook(xlsx_path)
    ws = wb.active
    
    # Get header row
    headers = [cell.value for cell in ws[1]]
    
    providers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not isinstance(row[0], str):
            continue
        
        provider = {}
        for i, header in enumerate(headers):
            if header and row[i] is not None:
                # Clean up string values
                value = str(row[i]).strip() if isinstance(row[i], str) else row[i]
                provider[header] = value
        
        if provider.get("Provider Name"):
            providers.append(provider)
    
    return providers


def determine_provider_type(provider: dict) -> str:
    """Determine provider type based on API compatibility."""
    openai_compat = provider.get("OpenAI Compatible", "").lower()
    anthropic_compat = provider.get("Anthropic Compatible", "").lower()
    api_type = provider.get("API Type", "").lower()
    
    if anthropic_compat in ("yes", "true", "y"):
        return "anthropic"
    elif openai_compat in ("yes", "true", "y", "yes (direct endpoint)"):
        return "openai"
    elif api_type in ("openai", "anthropic"):
        return api_type
    else:
        # Default to openai if unsure
        return "openai"


def generate_env_file(providers: list[dict], output_path: Path) -> None:
    """Generate .env file with API keys."""
    lines = [
        "# Auto-generated from providers.xlsx",
        "# Each key name must match the 'api_key_env' value used in providers.json",
        "",
    ]
    
    for provider in providers:
        name = provider.get("Provider Name", "")
        api_key = provider.get("API Key", "")
        
        if not name or not api_key:
            continue
        
        # Create environment variable name
        env_var = sanitize_name(name).upper() + "_API_KEY"
        lines.append(f"{env_var}={api_key}")
        lines.append("")
    
    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Generated .env file: {output_path}")


def generate_providers_json(providers: list[dict], output_path: Path) -> None:
    """Generate providers.json file."""
    config = {
        "providers": {},
        "test_prompt": "Reply OK only",
        "max_tokens": 10,
        "timeout_seconds": 30,
        "concurrency": 5
    }
    
    for provider in providers:
        name = provider.get("Provider Name", "")
        base_url = provider.get("Base URL", "")
        api_key = provider.get("API Key", "")
        
        if not name or not base_url:
            continue
        
        # Create provider key (lowercase, spaces to underscores)
        provider_key = sanitize_name(name).lower()
        
        # Determine type
        provider_type = determine_provider_type(provider)
        
        # Create environment variable name
        env_var = sanitize_name(name).upper() + "_API_KEY"
        
        provider_config = {
            "type": provider_type,
            "base_url": base_url.rstrip("/"),
            "api_key_env": env_var,
        }
        
        # Add optional fields
        if provider_type == "anthropic":
            provider_config["anthropic_version"] = provider.get("Anthropic Version", "2023-06-01")
        
        # Add known_models if available
        recommended_models = provider.get("Recommended Models", "")
        if recommended_models:
            models = [m.strip() for m in recommended_models.split(",") if m.strip()]
            provider_config["known_models"] = models
        
        # Add timeout if specified
        if provider.get("Rate Limit"):
            # Try to extract a reasonable timeout from rate limit info
            provider_config["timeout_seconds"] = 45
        
        # Add measure_ttft for high-priority providers
        priority = provider.get("Priority", "").lower()
        if priority in ("high", "recommended"):
            provider_config["measure_ttft"] = True
        
        config["providers"][provider_key] = provider_config
    
    output_path.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Generated providers.json: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate .env and providers.json from providers.xlsx")
    parser.add_argument("--xlsx", default="providers.xlsx", help="Path to providers.xlsx file")
    parser.add_argument("--env-file", default=".env", help="Output path for .env file")
    parser.add_argument("--providers", default="providers.json", help="Output path for providers.json")
    parser.add_argument("--output-dir", default=".", help="Output directory for generated files")
    args = parser.parse_args()
    
    # Resolve paths
    xlsx_path = Path(args.xlsx)
    output_dir = Path(args.output_dir)
    env_path = output_dir / args.env_file
    providers_path = output_dir / args.providers
    
    if not xlsx_path.exists():
        print(f"Error: Excel file not found: {xlsx_path}")
        return 1
    
    # Create output directory if needed
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Read providers from Excel
    print(f"Reading providers from: {xlsx_path}")
    providers = read_providers_from_excel(xlsx_path)
    print(f"Found {len(providers)} providers in Excel file")
    
    # Generate files
    generate_env_file(providers, env_path)
    generate_providers_json(providers, providers_path)
    
    print("\nDone! You can now run:")
    print(f"  python check_models.py --config {providers_path} --env {env_path}")
    
    return 0


if __name__ == "__main__":
    exit(main())
