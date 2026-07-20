#!/usr/bin/env python3
"""
ai-model-checker
=================

A general, provider-agnostic tool that checks which AI models are actually
usable (not just "listed") before you add them to something like an
opencode.json config.

Supported provider types today:
  - "openai"     : GET /models, POST /chat/completions, Authorization: Bearer <key>
  - "anthropic"  : POST /messages, headers x-api-key / anthropic-version

Per-provider optional fields in providers.json:
  - "known_models": [...]     fallback model list when /models isn't available
  - "timeout_seconds": N      overrides the global --timeout for this provider only
  - "measure_ttft": true      test via streaming and record time-to-first-token

Usage:
  python check_models.py                          # run full check
  python check_models.py --providers bluesminds    # only check one provider
  python check_models.py --generate-config out.json
  python check_models.py --concurrency 8 --timeout 20
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import concurrent.futures as futures
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import logging
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from dotenv import load_dotenv
try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl import load_workbook as openpyxl_load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    openpyxl = None
    OPENPYXL_AVAILABLE = False


# --------------------------------------------------------------------------
# Data models
# --------------------------------------------------------------------------

@dataclass
class ModelResult:
    provider: str
    model: str
    status: str            # "WORKING" | "FAILED"
    error: str = ""
    latency_ms: Optional[int] = None       # total round-trip time
    ttft_ms: Optional[int] = None          # time to first token (only when streaming test used)
    category: str = "general"


@dataclass
class ProviderConfig:
    name: str
    type: str
    base_url: str
    api_key: str
    anthropic_version: str = "2023-06-01"
    known_models: list[str] = field(default_factory=list)
    timeout_seconds: Optional[int] = None   # per-provider override of the global --timeout
    measure_ttft: bool = False              # if true, test via streaming and record time-to-first-token


# --------------------------------------------------------------------------
# Provider implementations
# --------------------------------------------------------------------------

class BaseProvider:
    """Common interface every provider type must implement."""

    def __init__(self, cfg: ProviderConfig, timeout: int, test_prompt: str, max_tokens: int):
        self.cfg = cfg
        # A provider-specific "timeout_seconds" in providers.json wins over
        # the global --timeout/default, since some routers are just slower.
        self.timeout = cfg.timeout_seconds if cfg.timeout_seconds is not None else timeout
        self.test_prompt = test_prompt
        self.max_tokens = max_tokens
        # Use a requests.Session per provider to reuse connections.
        self.session = requests.Session()
        # sensible default retry for network-level issues; provider-level 429 handling is done in logic
        retries = Retry(total=3, backoff_factor=0.5, status_forcelist=(502, 503, 504))
        adapter = HTTPAdapter(max_retries=retries)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

    def fetch_models(self) -> list[str]:
        raise NotImplementedError

    def test_model(self, model: str) -> ModelResult:
        raise NotImplementedError


class OpenAICompatibleProvider(BaseProvider):
    """Works with any provider implementing the OpenAI /v1 API shape
    (OpenAI itself, OpenRouter, BlueSminds, NaraRouter, etc.)."""

    def _headers(self) -> dict:
        return {
            "Authorization": f"Bearer {self.cfg.api_key}",
            "Content-Type": "application/json",
        }

    def fetch_models(self) -> list[str]:
        url = f"{self.cfg.base_url.rstrip('/')}/models"
        try:
            resp = self.session.get(url, headers=self._headers(), timeout=self.timeout)
            resp.raise_for_status()
            data = resp.json()
            # OpenAI-style: {"data": [{"id": "..."}]}
            items = data.get("data", data if isinstance(data, list) else [])
            model_ids = []
            for item in items:
                if isinstance(item, dict) and "id" in item:
                    model_ids.append(item["id"])
                elif isinstance(item, str):
                    model_ids.append(item)
            if model_ids:
                return model_ids
        except requests.RequestException as e:
            if not self.cfg.known_models:
                raise
            logger.warning("%s: /models fetch failed (%s), falling back to known_models", self.cfg.name, e)
        # Fall back to manually configured model list, if any.
        return list(self.cfg.known_models)

    def test_model(self, model: str) -> ModelResult:
        if self.cfg.measure_ttft:
            return self._test_model_streaming(model)
        return self._test_model_plain(model)

    def _test_model_plain(self, model: str) -> ModelResult:
        url = f"{self.cfg.base_url.rstrip('/')}/chat/completions"
        # Reasoning-family models (o1/o3/o4/gpt-5 style) reject "max_tokens"
        # and require "max_completion_tokens" instead - using the wrong key
        # causes a real, working model to be reported as FAILED.
        token_key = "max_completion_tokens" if classify_model(model) == "reasoning" else "max_tokens"
        payload = {
            "model": model,
            "messages": [{"role": "user", "content": self.test_prompt}],
            token_key: self.max_tokens,
        }
        start = time.time()
        # Retry once on 429 with short backoff; other transient network retries are handled by Session adapter.
        for attempt in range(2):
            try:
                resp = self.session.post(url, headers=self._headers(), json=payload, timeout=self.timeout)
                latency_ms = int((time.time() - start) * 1000)
                if resp.status_code == 429 and attempt == 0:
                    # Rate limited under concurrent testing, not a real failure - retry once.
                    time.sleep(2)
                    continue
                if resp.status_code != 200:
                    err = _extract_error(resp)
                    # Check if this might be a token parameter issue - retry with alternate key
                    if resp.status_code in (400, 422) or any(
                        param in err.lower() for param in ["max_tokens", "max_completion_tokens", "unknown"]
                    ):
                        alt_key = "max_completion_tokens" if token_key == "max_tokens" else "max_tokens"
                        alt_payload = {**payload, alt_key: self.max_tokens}
                        del alt_payload[token_key]
                        retry_start = time.time()
                        try:
                            retry_resp = self.session.post(url, headers=self._headers(), json=alt_payload, timeout=self.timeout)
                            retry_latency_ms = int((time.time() - retry_start) * 1000)
                            if retry_resp.status_code == 200:
                                retry_body = retry_resp.json()
                                if "choices" in retry_body and retry_body["choices"]:
                                    logger.info("%s: retried with alternate token key '%s' and succeeded for model %s", self.cfg.name, alt_key, model)
                                    return ModelResult(self.cfg.name, model, "WORKING", "", retry_latency_ms, category=classify_model(model))
                            # Retry failed - build combined error message
                            retry_err = _extract_error(retry_resp)
                            combined_err = f"first={err[:160]}; retry={retry_err[:160]}"
                            return ModelResult(self.cfg.name, model, "FAILED", combined_err, latency_ms, category=classify_model(model))
                        except requests.RequestException as e:
                            combined_err = f"first={err[:160]}; retry={str(e)[:160]}"
                            return ModelResult(self.cfg.name, model, "FAILED", combined_err, latency_ms, category=classify_model(model))
                    return ModelResult(self.cfg.name, model, "FAILED", err, latency_ms, category=classify_model(model))
                body = resp.json()
                if "choices" not in body or not body["choices"]:
                    return ModelResult(self.cfg.name, model, "FAILED", "no choices in response", latency_ms, category=classify_model(model))
                return ModelResult(self.cfg.name, model, "WORKING", "", latency_ms, category=classify_model(model))
            except requests.RequestException as e:
                latency_ms = int((time.time() - start) * 1000)
                return ModelResult(self.cfg.name, model, "FAILED", str(e), latency_ms, category=classify_model(model))
        # Exhausted retries (still 429)
        latency_ms = int((time.time() - start) * 1000)
        return ModelResult(self.cfg.name, model, "FAILED", "rate limited (429) after retry", latency_ms, category=classify_model(model))

    def _test_model_streaming(self, model: str) -> ModelResult:
        """Same test, but via stream=True so we can record time-to-first-token
        in addition to total latency. Falls back to a FAILED result (rather
        than silently switching to non-streaming) if the provider rejects
        streaming, so the user knows to disable measure_ttft for it."""
        url = f"{self.cfg.base_url.rstrip('/')}/chat/completions"
        token_key = "max_completion_tokens" if classify_model(model) == "reasoning" else "max_tokens"
        payload = {
            "model": model,
            "messages": [{"role": "user", "content": self.test_prompt}],
            token_key: self.max_tokens,
            "stream": True,
        }
        start = time.time()
        try:
            with self.session.post(url, headers=self._headers(), json=payload, timeout=self.timeout, stream=True) as resp:
                if resp.status_code != 200:
                    latency_ms = int((time.time() - start) * 1000)
                    err = _extract_error(resp)
                    return ModelResult(self.cfg.name, model, "FAILED", err, latency_ms, category=classify_model(model))

                ttft_ms = None
                got_any_token = False
                for raw_line in resp.iter_lines(decode_unicode=True):
                    if not raw_line or not raw_line.startswith("data:"):
                        continue
                    data_str = raw_line[len("data:"):].strip()
                    if data_str == "[DONE]":
                        break
                    try:
                        chunk = json.loads(data_str)
                    except json.JSONDecodeError:
                        continue
                    choices = chunk.get("choices") or []
                    delta = choices[0].get("delta", {}) if choices else {}
                    if delta.get("content") and ttft_ms is None:
                        ttft_ms = int((time.time() - start) * 1000)
                        got_any_token = True
                total_latency_ms = int((time.time() - start) * 1000)
                if not got_any_token:
                    return ModelResult(self.cfg.name, model, "FAILED", "no streamed content received", total_latency_ms, category=classify_model(model))
                return ModelResult(self.cfg.name, model, "WORKING", "", total_latency_ms, ttft_ms, classify_model(model))
        except requests.RequestException as e:
            latency_ms = int((time.time() - start) * 1000)
            return ModelResult(self.cfg.name, model, "FAILED", str(e), latency_ms, category=classify_model(model))


class AnthropicCompatibleProvider(BaseProvider):
    """Works with providers implementing the Anthropic /v1/messages API
    shape (Anthropic itself, AgentRouter, etc.)."""

    def _headers(self) -> dict:
        return {
            "x-api-key": self.cfg.api_key,
            "anthropic-version": self.cfg.anthropic_version,
            "Content-Type": "application/json",
        }

    def fetch_models(self) -> list[str]:
        # Not every Anthropic-compatible router exposes /models.
        # Try it; if it fails or returns nothing, fall back to a models
        # list supplied manually via "known_models" in providers.json.
        url = f"{self.cfg.base_url.rstrip('/')}/models"
        try:
            resp = self.session.get(url, headers=self._headers(), timeout=self.timeout)
            resp.raise_for_status()
            data = resp.json()
            items = data.get("data", data if isinstance(data, list) else [])
            model_ids = []
            for item in items:
                if isinstance(item, dict) and "id" in item:
                    model_ids.append(item["id"])
                elif isinstance(item, str):
                    model_ids.append(item)
            if model_ids:
                return model_ids
        except requests.RequestException:
            pass
        return list(self.cfg.known_models)

    def test_model(self, model: str) -> ModelResult:
        if self.cfg.measure_ttft:
            return self._test_model_streaming(model)
        return self._test_model_plain(model)

    def _test_model_plain(self, model: str) -> ModelResult:
        url = f"{self.cfg.base_url.rstrip('/')}/messages"
        payload = {
            "model": model,
            "max_tokens": self.max_tokens,
            "messages": [{"role": "user", "content": self.test_prompt}],
        }
        start = time.time()
        for attempt in range(2):
            try:
                resp = self.session.post(url, headers=self._headers(), json=payload, timeout=self.timeout)
                latency_ms = int((time.time() - start) * 1000)
                if resp.status_code == 429 and attempt == 0:
                    time.sleep(2)
                    continue
                if resp.status_code != 200:
                    err = _extract_error(resp)
                    return ModelResult(self.cfg.name, model, "FAILED", err, latency_ms, category=classify_model(model))
                body = resp.json()
                if "content" not in body or not body["content"]:
                    return ModelResult(self.cfg.name, model, "FAILED", "no content in response", latency_ms, category=classify_model(model))
                return ModelResult(self.cfg.name, model, "WORKING", "", latency_ms, category=classify_model(model))
            except requests.RequestException as e:
                latency_ms = int((time.time() - start) * 1000)
                return ModelResult(self.cfg.name, model, "FAILED", str(e), latency_ms, category=classify_model(model))
        latency_ms = int((time.time() - start) * 1000)
        return ModelResult(self.cfg.name, model, "FAILED", "rate limited (429) after retry", latency_ms, category=classify_model(model))

    def _test_model_streaming(self, model: str) -> ModelResult:
        """Streams the response so we can record time-to-first-token.
        Anthropic SSE sends "event: content_block_delta" lines for token
        deltas; the first one marks TTFT."""
        url = f"{self.cfg.base_url.rstrip('/')}/messages"
        payload = {
            "model": model,
            "max_tokens": self.max_tokens,
            "messages": [{"role": "user", "content": self.test_prompt}],
            "stream": True,
        }
        start = time.time()
        try:
            with self.session.post(url, headers=self._headers(), json=payload, timeout=self.timeout, stream=True) as resp:
                if resp.status_code != 200:
                    latency_ms = int((time.time() - start) * 1000)
                    err = _extract_error(resp)
                    return ModelResult(self.cfg.name, model, "FAILED", err, latency_ms, category=classify_model(model))

                ttft_ms = None
                got_any_token = False
                for raw_line in resp.iter_lines(decode_unicode=True):
                    if not raw_line or not raw_line.startswith("data:"):
                        continue
                    data_str = raw_line[len("data:"):].strip()
                    try:
                        chunk = json.loads(data_str)
                    except json.JSONDecodeError:
                        continue
                    if chunk.get("type") == "content_block_delta" and ttft_ms is None:
                        ttft_ms = int((time.time() - start) * 1000)
                        got_any_token = True
                    elif chunk.get("type") == "message_stop":
                        break
                total_latency_ms = int((time.time() - start) * 1000)
                if not got_any_token:
                    return ModelResult(self.cfg.name, model, "FAILED", "no streamed content received", total_latency_ms, category=classify_model(model))
                return ModelResult(self.cfg.name, model, "WORKING", "", total_latency_ms, ttft_ms, classify_model(model))
        except requests.RequestException as e:
            latency_ms = int((time.time() - start) * 1000)
            return ModelResult(self.cfg.name, model, "FAILED", str(e), latency_ms, category=classify_model(model))


PROVIDER_CLASSES = {
    "openai": OpenAICompatibleProvider,
    "anthropic": AnthropicCompatibleProvider,
}


def _extract_error(resp: requests.Response) -> str:
    try:
        data = resp.json()
        if isinstance(data, dict):
            err = data.get("error", data)
            if isinstance(err, dict):
                return err.get("message") or err.get("code") or str(err)
            return str(err)
    except ValueError:
        pass
    # Defensive: guard very long responses
    text = resp.text if isinstance(resp.text, str) else str(resp.text)
    return f"HTTP {resp.status_code}: {text[:200]}"


def classify_model(model_name: str) -> str:
    """Best-effort category classification from the model name alone."""
    name = model_name.lower()
    if any(k in name for k in ("code", "coder", "codex")):
        return "coding"
    if any(k in name for k in ("vision", "-vl", "vl-", "visual")):
        return "vision"
    if any(k in name for k in ("o1", "o3", "o4", "reason", "think")):
        return "reasoning"
    if any(k in name for k in ("mini", "flash", "haiku", "nano", "small", "lite")):
        return "fast"
    return "general"


# Configure basic logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)


# --------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------

def load_provider_configs(config_path: Path, only: Optional[list[str]]) -> list[ProviderConfig]:
    raw = json.loads(config_path.read_text(encoding="utf-8"))
    configs = []
    for name, entry in raw.get("providers", {}).items():
        if only and name not in only:
            continue
        api_key_env = entry.get("api_key_env")
        api_key = os.environ.get(api_key_env, "") if api_key_env else ""
        if not api_key:
            logger.warning("no API key found for provider '%s' (env var '%s') - skipping", name, api_key_env)
            continue
        configs.append(ProviderConfig(
            name=name,
            type=entry.get("type", "openai"),
            base_url=entry["base_url"],
            api_key=api_key,
            anthropic_version=entry.get("anthropic_version", "2023-06-01"),
            known_models=entry.get("known_models", []),
            timeout_seconds=entry.get("timeout_seconds"),
            measure_ttft=entry.get("measure_ttft", False),
        ))
    return configs


def build_provider(cfg: ProviderConfig, timeout: int, test_prompt: str, max_tokens: int) -> BaseProvider:
    cls = PROVIDER_CLASSES.get(cfg.type)
    if cls is None:
        raise ValueError(f"Unknown provider type '{cfg.type}' for provider '{cfg.name}'")
    return cls(cfg, timeout, test_prompt, max_tokens)


def run_checks(providers: list[BaseProvider], concurrency: int) -> list[ModelResult]:
    results: list[ModelResult] = []
    for provider in providers:
        logger.info("=== %s (%s) ===", provider.cfg.name, provider.cfg.type)
        try:
            models = provider.fetch_models()
        except requests.RequestException as e:
            logger.error("could not fetch model list for %s: %s", provider.cfg.name, e)
            continue

        if not models:
            logger.warning("no models returned/found for provider %s", provider.cfg.name)
            continue

        logger.info("found %d model(s), testing with max %d in parallel...", len(models), concurrency)

        with futures.ThreadPoolExecutor(max_workers=concurrency) as pool:
            future_map = {pool.submit(provider.test_model, m): m for m in models}
            for fut in futures.as_completed(future_map):
                result = fut.result()
                results.append(result)
                mark = "OK" if result.status == "WORKING" else "FAIL"
                latency = f"{result.latency_ms}ms" if result.latency_ms is not None else "-"
                ttft = f" ttft={result.ttft_ms}ms" if result.ttft_ms is not None else ""
                logger.info("  [%s] %s %8s%s  %s", mark, result.model, latency, ttft, result.error)

    return results


def write_csv(results: list[ModelResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["provider", "model", "status", "latency_ms", "ttft_ms", "category", "error"])
        for r in results:
            writer.writerow([r.provider, r.model, r.status, r.latency_ms or "", r.ttft_ms or "", r.category, r.error])
    logger.info("report written to %s", output_path)
    # Also write an Excel file with two sheets: OK and FAILED
    if openpyxl is None:
        logger.info("openpyxl not installed; skipping XLSX output")
        return
    try:
        wb = Workbook()
        ws_ok = wb.active
        ws_ok.title = "OK"
        ws_fail = wb.create_sheet(title="FAILED")

        header = ["provider", "model", "status", "latency_ms", "ttft_ms", "category", "error"]
        ws_ok.append(header)
        ws_fail.append(header)

        for r in results:
            row = [r.provider, r.model, r.status, r.latency_ms or "", r.ttft_ms or "", r.category, r.error]
            if r.status == "WORKING":
                ws_ok.append(row)
            else:
                ws_fail.append(row)

        xlsx_path = output_path.with_suffix('.xlsx')
        wb.save(xlsx_path)
        logger.info("Excel report written to %s", xlsx_path)
    except Exception as e:
        logger.error("failed to write Excel report: %s", e)


def generate_opencode_config(results: list[ModelResult], output_path: Path) -> None:
    config: dict = {"provider": {}}
    for r in results:
        if r.status != "WORKING":
            continue
        provider_block = config["provider"].setdefault(r.provider, {"models": {}})
        provider_block["models"][r.model] = {"name": r.model}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
    working_count = sum(1 for r in results if r.status == "WORKING")
    logger.info("opencode config with %d working model(s) written to %s", working_count, output_path)


def merge_opencode_config(results: list[ModelResult], real_config_path: str = None) -> None:
    """Merge working models into the real opencode config file."""
    if real_config_path is None:
        # Default path for opencode config
        real_config_path = os.path.expanduser("~/.config/opencode/opencode.jsonc")
    
    real_config_path = Path(real_config_path)
    
    if not real_config_path.exists():
        logger.warning("Real opencode config not found at %s, skipping merge", real_config_path)
        return
    
    # Load existing config
    try:
        with open(real_config_path, 'r', encoding='utf-8') as f:
            real_config = json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        logger.error("Failed to read real opencode config: %s", e)
        return
    
    # Build new provider entries from working models
    new_providers = {}
    for r in results:
        if r.status != "WORKING":
            continue
        if r.provider not in new_providers:
            new_providers[r.provider] = {"models": {}}
        new_providers[r.provider]["models"][r.model] = {"name": r.model}
    
    # Merge new providers into existing config
    if "provider" not in real_config:
        real_config["provider"] = {}
    
    for provider_name, provider_data in new_providers.items():
        if provider_name not in real_config["provider"]:
            real_config["provider"][provider_name] = provider_data
        else:
            # Merge models
            existing_models = real_config["provider"][provider_name].get("models", {})
            for model_name, model_config in provider_data["models"].items():
                if model_name not in existing_models:
                    existing_models[model_name] = model_config
    
    # Write back the merged config
    try:
        with open(real_config_path, 'w', encoding='utf-8') as f:
            json.dump(real_config, f, indent=2, ensure_ascii=False)
        working_count = sum(1 for r in results if r.status == "WORKING")
        logger.info("Merged %d working model(s) into real opencode config at %s", working_count, real_config_path)
    except IOError as e:
        logger.error("Failed to write merged opencode config: %s", e)


def read_providers_from_excel(xlsx_path: Path) -> list[dict]:
    """Read provider data from Excel file."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required to read Excel files. Install with: pip install openpyxl")
    
    wb = openpyxl_load_workbook(xlsx_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    providers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not isinstance(row[0], str):
            continue
        
        provider = {}
        for i, header in enumerate(headers):
            if header and row[i] is not None:
                value = str(row[i]).strip() if isinstance(row[i], str) else row[i]
                provider[header] = value
        
        if provider.get("Provider Name"):
            providers.append(provider)
    
    return providers


def determine_provider_type(provider: dict) -> str:
    """Determine provider type based on API compatibility."""
    openai_compat = provider.get("OpenAI Compatible", "").lower()
    anthropic_compat = provider.get("Anthropic Compatible", "").lower()
    api_type = provider.get("API Type", "").lower()
    
    if anthropic_compat in ("yes", "true", "y"):
        return "anthropic"
    elif openai_compat in ("yes", "true", "y", "yes (direct endpoint)"):
        return "openai"
    elif api_type in ("openai", "anthropic"):
        return api_type
    else:
        return "openai"


def sanitize_name(name: str) -> str:
    """Sanitize provider name for use as environment variable or JSON key."""
    import re
    # Remove parentheses and their contents
    name = re.sub(r'\(.*?\)', '', name)
    # Remove special characters and replace spaces with underscores
    name = re.sub(r'[^a-zA-Z0-9\s\-]', '', name)
    # Replace spaces and hyphens with underscores
    name = name.replace(' ', '_').replace('-', '_')
    # Replace multiple consecutive underscores with single underscore
    name = re.sub(r'_+', '_', name)
    # Remove leading/trailing underscores
    name = name.strip('_')
    return name


def generate_env_file(providers: list[dict], output_path: Path) -> None:
    """Generate .env file with API keys."""
    lines = [
        "# Auto-generated from providers.xlsx",
        "# Each key name must match the 'api_key_env' value used in providers.json",
        "",
    ]
    
    for provider in providers:
        name = provider.get("Provider Name", "")
        api_key = provider.get("API Key", "")
        
        if not name or not api_key:
            continue
        
        env_var = sanitize_name(name).upper() + "_API_KEY"
        lines.append(f"{env_var}={api_key}")
        lines.append("")
    
    output_path.write_text("\n".join(lines), encoding="utf-8")


def generate_providers_json(providers: list[dict], output_path: Path) -> None:
    """Generate providers.json file."""
    config = {
        "providers": {},
        "test_prompt": "Reply OK only",
        "max_tokens": 10,
        "timeout_seconds": 30,
        "concurrency": 5
    }
    
    for provider in providers:
        name = provider.get("Provider Name", "")
        base_url = provider.get("Base URL", "")
        
        if not name or not base_url:
            continue
        
        provider_key = sanitize_name(name).lower()
        provider_type = determine_provider_type(provider)
        env_var = sanitize_name(name).upper() + "_API_KEY"
        
        provider_config = {
            "type": provider_type,
            "base_url": base_url.rstrip("/"),
            "api_key_env": env_var,
        }
        
        if provider_type == "anthropic":
            provider_config["anthropic_version"] = provider.get("Anthropic Version", "2023-06-01")
        
        recommended_models = provider.get("Recommended Models", "")
        if recommended_models:
            models = [m.strip() for m in recommended_models.split(",") if m.strip()]
            provider_config["known_models"] = models
        
        priority = provider.get("Priority", "").lower()
        if priority in ("high", "recommended"):
            provider_config["measure_ttft"] = True
        
        config["providers"][provider_key] = provider_config
    
    output_path.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")


def cmd_generate_config(args) -> int:
    """Generate .env and providers.json from Excel file."""
    xlsx_path = Path(args.xlsx)
    output_dir = Path(args.output_dir) if args.output_dir else Path(".")
    env_path = output_dir / (args.env_file or ".env")
    providers_path = output_dir / (args.providers or "providers.json")
    
    if not xlsx_path.exists():
        logger.error("Excel file not found: %s", xlsx_path)
        return 1
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"Reading providers from: {xlsx_path}")
    providers = read_providers_from_excel(xlsx_path)
    print(f"Found {len(providers)} providers in Excel file")
    
    generate_env_file(providers, env_path)
    generate_providers_json(providers, providers_path)
    
    print(f"\nGenerated: {env_path}")
    print(f"Generated: {providers_path}")
    print("\nYou can now run:")
    print(f"  python check_models.py --config {providers_path} --env {env_path}")
    
    return 0


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Check which AI provider models actually work.")
    subparsers = parser.add_subparsers(dest="command", help="Available commands")
    
    # Main check command
    check_parser = subparsers.add_parser("check", help="Check which AI provider models actually work")
    check_parser.add_argument("--config", default="providers.json", help="path to providers.json")
    check_parser.add_argument("--env", default=".env", help="path to .env file with API keys")
    check_parser.add_argument("--providers", nargs="*", default=None, help="only check these provider names")
    check_parser.add_argument("--output", default="results/models_report.csv", help="CSV report output path")
    check_parser.add_argument("--generate-config", default=None, help="also write an opencode-style JSON config of working models")
    check_parser.add_argument("--merge-opencode", action="store_true", help="merge working models into the real opencode config (~/.config/opencode/opencode.jsonc)")
    check_parser.add_argument("--opencode-path", default=None, help="custom path to opencode config file (default: ~/.config/opencode/opencode.jsonc)")
    check_parser.add_argument("--concurrency", type=int, default=None, help="override concurrency from providers.json")
    check_parser.add_argument("--timeout", type=int, default=None, help="override per-request timeout (seconds)")
    check_parser.add_argument("--auto-update", action="store_true", help="auto-update .env and providers.json from providers.xlsx before running checks")
    check_parser.add_argument("--xlsx", default="providers.xlsx", help="path to providers.xlsx file (used with --auto-update)")
    
    # Generate config command
    gen_parser = subparsers.add_parser("generate", help="Generate .env and providers.json from Excel file")
    gen_parser.add_argument("--xlsx", default="providers.xlsx", help="path to providers.xlsx file")
    gen_parser.add_argument("--env-file", default=".env", help="output path for .env file")
    gen_parser.add_argument("--providers", default="providers.json", help="output path for providers.json")
    gen_parser.add_argument("--output-dir", default=".", help="output directory for generated files")
    
    args = parser.parse_args()
    
    if args.command == "generate":
        sys.exit(cmd_generate_config(args))
    
    # Default to check command if no subcommand specified
    if args.command is None:
        args = check_parser.parse_args([])
    
    # Auto-update from Excel if requested
    if getattr(args, 'auto_update', False) and OPENPYXL_AVAILABLE:
        xlsx_path = Path(getattr(args, 'xlsx', 'providers.xlsx'))
        if xlsx_path.exists():
            print(f"Auto-updating config from: {xlsx_path}")
            providers_data = read_providers_from_excel(xlsx_path)
            generate_env_file(providers_data, Path(args.env))
            generate_providers_json(providers_data, Path(args.config))
            print(f"Updated: {args.env}")
            print(f"Updated: {args.config}")
            # Reload environment variables from the updated file
            load_dotenv(args.env, override=True)
        else:
            logger.warning(f"Excel file not found: {xlsx_path}, using existing config files")
            load_dotenv(args.env)
    else:
        load_dotenv(args.env)

    config_path = Path(args.config)
    if not config_path.exists():
        logger.error("config file not found: %s", config_path)
        sys.exit(1)

    raw_cfg = json.loads(config_path.read_text(encoding="utf-8"))
    test_prompt = raw_cfg.get("test_prompt", "Reply OK only")
    max_tokens = raw_cfg.get("max_tokens", 10)
    timeout = args.timeout or raw_cfg.get("timeout_seconds", 30)
    concurrency = args.concurrency or raw_cfg.get("concurrency", 5)

    provider_configs = load_provider_configs(config_path, args.providers)
    if not provider_configs:
        logger.error("no providers with valid API keys found - check your .env file")
        sys.exit(1)

    providers = [build_provider(c, timeout, test_prompt, max_tokens) for c in provider_configs]

    results = run_checks(providers, concurrency)

    if not results:
        logger.warning("no results collected")
        sys.exit(0)

    write_csv(results, Path(args.output))

    if args.generate_config:
        generate_opencode_config(results, Path(args.generate_config))
    
    if getattr(args, 'merge_opencode', False):
        opencode_path = getattr(args, 'opencode_path', None)
        merge_opencode_config(results, opencode_path)

    working = sum(1 for r in results if r.status == "WORKING")
    logger.info("Summary: %d/%d models working", working, len(results))


if __name__ == "__main__":
    main()
